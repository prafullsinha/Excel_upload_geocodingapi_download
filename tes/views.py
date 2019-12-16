import requests
from django.shortcuts import render, redirect, HttpResponse, Http404
import openpyxl
from .models import ExcelModel
from .forms import RegisterForm
from django.views.generic import TemplateView
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.contrib.auth import login
import mimetypes
from django.conf import settings
import os


class IndexView(TemplateView):
    template_name = 'tes/index.html'

    def get(self, request, *args, **kwargs):
        form = ExcelModel.objects.all()
        context = {
            'form': form
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]
        excel_data = list()
        count = 0
        i = 0
        for row in worksheet.iter_rows():
            i = i + 1
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            if count == 0:
                worksheet.cell(row=i, column=2).value = "Latitude"
                worksheet.cell(row=i, column=3).value = "Longitude"
                count = count + 1
            else:
                api_key = "AIzaSyCZwboOYJ3h6qbHv1QUEcT-SN-yU2TPOlI"
                address = row_data
                api_response = requests.get(
                    'https://maps.googleapis.com/maps/api/geocode/json?address={0}&key={1}'.format(address, api_key))
                api_response_dict = api_response.json()
                if api_response_dict['status'] == 'OK':
                    latitude = api_response_dict['results'][0]['geometry']['location']['lat']
                    longitude = api_response_dict['results'][0]['geometry']['location']['lng']
                    worksheet.cell(row=i, column=2).value = str(latitude)
                    worksheet.cell(row=i, column=3).value = str(longitude)
                    wb.save(excel_file)
        form = ExcelModel()
        form.excelfile = excel_file
        form.user = request.user
        form.save()
        q = ExcelModel.objects.filter(user=request.user)
        context = {
            'excel_data': excel_data,
            'q': q
        }
        return render(request, self.template_name, context)


class SignupView(TemplateView):
    template_name = 'registration/signup.html'

    def get(self, request, *args, **kwargs):
        form = RegisterForm()
        context = {
            'form': form
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = User.objects.create(username=form.cleaned_data['email'],
                                       first_name=form.cleaned_data['first_name'],
                                       password=make_password(form.cleaned_data['password1']))
            login(request, user)
            return redirect('tes:index')
        context = {
            'form': form
        }
        return render(request, self.template_name, context)


def download_file(request, filepath):
    file_path = os.path.join(settings.MEDIA_ROOT, filepath)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404
